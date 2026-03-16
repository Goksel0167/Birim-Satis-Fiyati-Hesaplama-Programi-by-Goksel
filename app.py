"""
Birim Satış Fiyatı Hesaplayıcı - Flask Uygulaması
Formül: ((NTS Maliyeti / Marj) + Nakliye) / Döviz Kurları
TCMB döviz kurları ile otomatik hesaplama
"""

import os
import json
import uuid
import sqlite3
from datetime import datetime, timedelta
from io import BytesIO
from contextlib import contextmanager

import requests
import xml.etree.ElementTree as ET
from flask import Flask, render_template, request, jsonify, send_file

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'birim-fiyat-hesaplayici-secret')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get('DATABASE_PATH', os.path.join(BASE_DIR, 'data', 'calculations.db'))

try:
    MAX_RECORDS = max(1, int(os.environ.get('MAX_RECORDS', '500')))
except ValueError:
    MAX_RECORDS = 500

# ─── DATABASE ──────────────────────────────────────────────────────

def get_db():
    """Get a database connection."""
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize the database table."""
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS calculations (
            id TEXT PRIMARY KEY,
            product_name TEXT NOT NULL,
            dealer TEXT NOT NULL,
            dealer_customer TEXT NOT NULL,
            calculation_date TEXT NOT NULL,
            usd_rate REAL NOT NULL,
            eur_rate REAL NOT NULL,
            chf_rate REAL NOT NULL,
            factory TEXT NOT NULL,
            shipping_city TEXT NOT NULL DEFAULT '',
            shipping_cost REAL NOT NULL,
            shipping_cost_usd REAL NOT NULL DEFAULT 0,
            nts_cost REAL NOT NULL,
            margin REAL NOT NULL,
            result_tl REAL NOT NULL,
            result_usd REAL NOT NULL,
            result_eur REAL NOT NULL,
            result_chf REAL NOT NULL,
            result_tl_ton REAL NOT NULL DEFAULT 0,
            result_usd_ton REAL NOT NULL DEFAULT 0,
            result_eur_ton REAL NOT NULL DEFAULT 0,
            result_chf_ton REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

def enforce_max_records():
    """Keep only the latest MAX_RECORDS records."""
    conn = get_db()
    count = conn.execute('SELECT COUNT(*) FROM calculations').fetchone()[0]
    if count > MAX_RECORDS:
        excess = count - MAX_RECORDS
        conn.execute('''
            DELETE FROM calculations WHERE id IN (
                SELECT id FROM calculations ORDER BY created_at ASC LIMIT ?
            )
        ''', (excess,))
        conn.commit()
    conn.close()

# ─── TCMB API ─────────────────────────────────────────────────────

def parse_tcmb_xml(xml_text):
    """Parse TCMB XML and extract USD, EUR, CHF selling rates."""
    root = ET.fromstring(xml_text)
    date_str = root.attrib.get('Tarih', '')
    
    rates = {'usd': 0, 'eur': 0, 'chf': 0, 'date': date_str}
    
    for currency in root.findall('Currency'):
        code = currency.attrib.get('CurrencyCode', '')
        forex_selling = currency.find('ForexSelling')
        if forex_selling is not None and forex_selling.text:
            value = float(forex_selling.text)
            if code == 'USD':
                rates['usd'] = value
            elif code == 'EUR':
                rates['eur'] = value
            elif code == 'CHF':
                rates['chf'] = value
    
    return rates

def build_tcmb_url(d):
    """Build TCMB URL for a given date."""
    day = str(d.day).zfill(2)
    month = str(d.month).zfill(2)
    year = str(d.year)
    return f"https://www.tcmb.gov.tr/kurlar/{year}{month}/{day}{month}{year}.xml"

def fetch_rates_with_fallback(target_date, max_retries=7):
    """Try fetching TCMB rates, going back up to max_retries days for weekends/holidays."""
    d = target_date
    for _ in range(max_retries):
        url = build_tcmb_url(d)
        try:
            resp = requests.get(url, timeout=10, headers={
                'User-Agent': 'Mozilla/5.0'
            })
            if resp.status_code == 200:
                return parse_tcmb_xml(resp.text)
        except Exception:
            pass
        d = d - timedelta(days=1)
    return None

def parse_date_str(date_str):
    """Parse DD-MM-YYYY string into a date object."""
    try:
        parts = date_str.split('-')
        if len(parts) == 3:
            return datetime(int(parts[2]), int(parts[1]), int(parts[0])).date()
    except (ValueError, IndexError):
        pass
    return None

# ─── ROUTES ────────────────────────────────────────────────────────

@app.route('/')
def index():
    """Main page."""
    return render_template('index.html')

@app.route('/api/exchange-rates/today')
def get_today_rates():
    """Get exchange rates for today (with fallback)."""
    try:
        # Try today.xml first
        resp = requests.get('https://www.tcmb.gov.tr/kurlar/today.xml', timeout=10, headers={
            'User-Agent': 'Mozilla/5.0'
        })
        if resp.status_code == 200:
            rates = parse_tcmb_xml(resp.text)
            return jsonify(rates)
        
        # Fallback
        rates = fetch_rates_with_fallback(datetime.now().date())
        if rates:
            return jsonify(rates)
        
        return jsonify({'error': 'TCMB API yanıt vermedi'}), 502
    except Exception as e:
        return jsonify({'error': f'Döviz kurları alınamadı: {str(e)}'}), 500

@app.route('/api/exchange-rates/<date_str>')
def get_date_rates(date_str):
    """Get exchange rates for a specific date (DD-MM-YYYY)."""
    try:
        target_date = parse_date_str(date_str)
        if not target_date:
            return jsonify({'error': 'Geçersiz tarih formatı. DD-MM-YYYY kullanın.'}), 400
        
        rates = fetch_rates_with_fallback(target_date)
        if rates:
            return jsonify(rates)
        
        return jsonify({'error': 'Bu tarih aralığında kur bulunamadı.'}), 404
    except Exception as e:
        return jsonify({'error': f'Döviz kurları alınamadı: {str(e)}'}), 500

@app.route('/api/calculations', methods=['GET'])
def get_calculations():
    """Get all calculations."""
    conn = get_db()
    rows = conn.execute('SELECT * FROM calculations ORDER BY created_at DESC').fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])

@app.route('/api/calculations', methods=['POST'])
def create_calculation():
    """Create a new calculation."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Geçersiz veri'}), 400
    
    calc_id = str(uuid.uuid4())
    created_at = datetime.now().isoformat()
    
    # Extract values
    nts_cost = float(data.get('nts_cost', 0))
    margin = float(data.get('margin', 0.70))
    shipping_cost = float(data.get('shipping_cost', 0))  # TL/kg
    usd_rate = float(data.get('usd_rate', 0))
    eur_rate = float(data.get('eur_rate', 0))
    chf_rate = float(data.get('chf_rate', 0))
    
    # Calculate shipping cost in USD/kg
    shipping_cost_usd = shipping_cost / usd_rate if usd_rate > 0 else 0
    
    # Calculate: (NTS / Margin + Shipping)
    base_tl = (nts_cost / margin) + shipping_cost  # TL/kg
    
    # Per-kg results
    result_tl = base_tl
    result_usd = base_tl / usd_rate if usd_rate > 0 else 0
    result_eur = base_tl / eur_rate if eur_rate > 0 else 0
    result_chf = base_tl / chf_rate if chf_rate > 0 else 0
    
    # Per-ton results (kg × 1000)
    result_tl_ton = result_tl * 1000
    result_usd_ton = result_usd * 1000
    result_eur_ton = result_eur * 1000
    result_chf_ton = result_chf * 1000
    
    conn = get_db()
    conn.execute('''
        INSERT INTO calculations (
            id, product_name, dealer, dealer_customer, calculation_date,
            usd_rate, eur_rate, chf_rate, factory, shipping_city,
            shipping_cost, shipping_cost_usd, nts_cost, margin,
            result_tl, result_usd, result_eur, result_chf,
            result_tl_ton, result_usd_ton, result_eur_ton, result_chf_ton,
            created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        calc_id,
        data.get('product_name', ''),
        data.get('dealer', ''),
        data.get('dealer_customer', ''),
        data.get('calculation_date', ''),
        usd_rate, eur_rate, chf_rate,
        data.get('factory', ''),
        data.get('shipping_city', ''),
        shipping_cost, shipping_cost_usd,
        nts_cost, margin,
        result_tl, result_usd, result_eur, result_chf,
        result_tl_ton, result_usd_ton, result_eur_ton, result_chf_ton,
        created_at
    ))
    conn.commit()
    
    # Enforce max records
    enforce_max_records()
    
    # Return the created record
    row = conn.execute('SELECT * FROM calculations WHERE id = ?', (calc_id,)).fetchone()
    conn.close()
    
    return jsonify(dict(row)), 201

@app.route('/api/calculations/<calc_id>', methods=['DELETE'])
def delete_calculation(calc_id):
    """Delete a calculation."""
    conn = get_db()
    cursor = conn.execute('DELETE FROM calculations WHERE id = ?', (calc_id,))
    conn.commit()
    conn.close()
    
    if cursor.rowcount == 0:
        return jsonify({'error': 'Hesaplama bulunamadı'}), 404
    
    return jsonify({'success': True})

# ─── EXPORT ROUTES ─────────────────────────────────────────────────

@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    """Export calculations to PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    data = request.get_json()
    calculations = data.get('calculations', [])
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           leftMargin=10*mm, rightMargin=10*mm,
                           topMargin=10*mm, bottomMargin=10*mm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=14, textColor=colors.HexColor('#296291')
    )
    sub_style = ParagraphStyle(
        'SubTitle', parent=styles['Normal'],
        fontSize=8, textColor=colors.grey
    )
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'],
        fontSize=6.5, leading=8
    )
    
    elements = []
    elements.append(Paragraph('Birim Satış Fiyatı Hesaplama Geçmişi', title_style))
    now_str = datetime.now().strftime('%d.%m.%Y %H:%M')
    elements.append(Paragraph(f'Oluşturulma: {now_str}  |  Toplam: {len(calculations)} hesaplama', sub_style))
    elements.append(Spacer(1, 4*mm))
    
    # Table header
    header = ['Tarih', 'Ürün', 'Bayi', 'Fab.', 'Şehir', 'Marj',
              'Nakl. TL/kg', 'USD/kg', 'EUR/kg', 'CHF/kg',
              'USD/ton', 'EUR/ton', 'CHF/ton']
    
    table_data = [[Paragraph(h, cell_style) for h in header]]
    
    for c in calculations:
        result_tl_ton = c.get('result_tl_ton', c.get('result_tl', 0) * 1000)
        result_usd_ton = c.get('result_usd_ton', c.get('result_usd', 0) * 1000)
        result_eur_ton = c.get('result_eur_ton', c.get('result_eur', 0) * 1000)
        result_chf_ton = c.get('result_chf_ton', c.get('result_chf', 0) * 1000)
        
        factory_labels = {'adana': 'Adana', 'trabzon': 'Trab.', 'gebze': 'Gebze'}
        
        row = [
            Paragraph(c.get('created_at', '')[:16].replace('T', ' '), cell_style),
            Paragraph(str(c.get('product_name', '')), cell_style),
            Paragraph(str(c.get('dealer', '')), cell_style),
            Paragraph(factory_labels.get(c.get('factory', ''), c.get('factory', '')), cell_style),
            Paragraph(str(c.get('shipping_city', ''))[:10], cell_style),
            Paragraph(f"%{int(c.get('margin', 0.7) * 100)}", cell_style),
            Paragraph(f"{c.get('shipping_cost', 0):.2f}", cell_style),
            Paragraph(f"{c.get('result_usd', 0):.4f}", cell_style),
            Paragraph(f"{c.get('result_eur', 0):.4f}", cell_style),
            Paragraph(f"{c.get('result_chf', 0):.4f}", cell_style),
            Paragraph(f"{result_usd_ton:.2f}", cell_style),
            Paragraph(f"{result_eur_ton:.2f}", cell_style),
            Paragraph(f"{result_chf_ton:.2f}", cell_style),
        ]
        table_data.append(row)
    
    col_widths = [35*mm, 30*mm, 25*mm, 14*mm, 20*mm, 12*mm,
                  20*mm, 20*mm, 20*mm, 20*mm, 22*mm, 22*mm, 22*mm]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#296291')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),
        ('ALIGN', (5, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    filename = f"hesaplamalar_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

@app.route('/api/export/pdf/single', methods=['POST'])
def export_single_pdf():
    """Export a single calculation to PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    calc = request.get_json()
    if not calc:
        return jsonify({'error': 'Geçersiz veri'}), 400
    
    factory_labels = {'adana': 'Adana', 'trabzon': 'Trabzon', 'gebze': 'Gebze'}
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           leftMargin=14*mm, rightMargin=14*mm,
                           topMargin=10*mm, bottomMargin=10*mm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontSize=16, textColor=colors.white,
        spaceAfter=2
    )
    section_style = ParagraphStyle(
        'Section', parent=styles['Heading2'],
        fontSize=11, textColor=colors.HexColor('#296291'),
        spaceBefore=8, spaceAfter=4
    )
    label_style = ParagraphStyle(
        'Label', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#505050'),
        fontName='Helvetica-Bold'
    )
    value_style = ParagraphStyle(
        'Value', parent=styles['Normal'],
        fontSize=9
    )
    big_style = ParagraphStyle(
        'Big', parent=styles['Normal'],
        fontSize=16, fontName='Helvetica-Bold'
    )
    small_style = ParagraphStyle(
        'Small', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#3c3c3c')
    )
    footer_style = ParagraphStyle(
        'Footer', parent=styles['Normal'],
        fontSize=7, textColor=colors.grey
    )
    
    elements = []
    
    # Header
    header_data = [[Paragraph('Birim Satış Fiyatı Hesaplama', title_style)]]
    header_table = Table(header_data, colWidths=[182*mm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#296291')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 6*mm))
    
    # Product info
    elements.append(Paragraph('Ürün ve Müşteri Bilgileri', section_style))
    info_data = [
        ['Ürün İsmi', calc.get('product_name', '')],
        ['Bayi', calc.get('dealer', '')],
        ['Bayi Müşteri', calc.get('dealer_customer', '')],
        ['Sevk Fabrikası', factory_labels.get(calc.get('factory', ''), calc.get('factory', ''))],
        ['Nakliye Şehri', calc.get('shipping_city', '') or '—'],
    ]
    info_table = Table(info_data, colWidths=[45*mm, 130*mm])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#505050')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 4*mm))
    
    # Exchange rates
    elements.append(Paragraph('Döviz Kurları (TCMB Satış)', section_style))
    rates_data = [
        ['Hesaplama Tarihi', calc.get('calculation_date', '')],
        ['USD Kur', f"{calc.get('usd_rate', 0):.4f} TL"],
        ['EUR Kur', f"{calc.get('eur_rate', 0):.4f} TL"],
        ['CHF Kur', f"{calc.get('chf_rate', 0):.4f} TL"],
    ]
    rates_table = Table(rates_data, colWidths=[45*mm, 130*mm])
    rates_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#505050')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
    ]))
    elements.append(rates_table)
    elements.append(Spacer(1, 4*mm))
    
    # Calculation details
    elements.append(Paragraph('Maliyet ve Hesaplama', section_style))
    nts_cost = calc.get('nts_cost', 0)
    margin_val = calc.get('margin', 0.7)
    shipping_cost_usd = calc.get('shipping_cost_usd', 0)
    if shipping_cost_usd == 0 and calc.get('usd_rate', 0) > 0:
        shipping_cost_usd = calc.get('shipping_cost', 0) / calc.get('usd_rate', 1)
    
    calc_data = [
        ['NTS Maliyeti', f"{nts_cost:.2f} TL/kg"],
        ['Nakliye Bedeli (TL)', f"{calc.get('shipping_cost', 0):.2f} TL/kg"],
        ['Nakliye Bedeli (USD)', f"{shipping_cost_usd:.4f} USD/kg"],
        ['Hedeflenen Marj', f"%{int(margin_val * 100)}"],
        ['NTS / Marj', f"{(nts_cost / margin_val if margin_val > 0 else 0):.2f} TL/kg"],
    ]
    calc_table = Table(calc_data, colWidths=[45*mm, 130*mm])
    calc_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#505050')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
    ]))
    elements.append(calc_table)
    elements.append(Spacer(1, 3*mm))
    
    elements.append(Paragraph('Formül: (NTS Maliyeti / Marj) + Nakliye = TL/kg Birim Fiyat', footer_style))
    elements.append(Spacer(1, 5*mm))
    
    # Results - KG
    result_tl = calc.get('result_tl', 0)
    result_usd = calc.get('result_usd', 0)
    result_eur = calc.get('result_eur', 0)
    result_chf = calc.get('result_chf', 0)
    
    kg_data = [
        [Paragraph('Birim Satış Fiyatı (kg bazında)', section_style)],
        [Paragraph(f"{result_tl:.2f} TL/kg", big_style)],
        [Paragraph(f"USD: {result_usd:.4f}/kg  |  EUR: {result_eur:.4f}/kg  |  CHF: {result_chf:.4f}/kg", small_style)],
    ]
    kg_table = Table(kg_data, colWidths=[170*mm])
    kg_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f5fa')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#296291')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(kg_table)
    elements.append(Spacer(1, 4*mm))
    
    # Results - TON
    result_tl_ton = calc.get('result_tl_ton', result_tl * 1000)
    result_usd_ton = calc.get('result_usd_ton', result_usd * 1000)
    result_eur_ton = calc.get('result_eur_ton', result_eur * 1000)
    result_chf_ton = calc.get('result_chf_ton', result_chf * 1000)
    
    ton_data = [
        [Paragraph('Birim Satış Fiyatı (ton bazında)', section_style)],
        [Paragraph(f"{result_tl_ton:.2f} TL/ton", big_style)],
        [Paragraph(f"USD: {result_usd_ton:.2f}/ton  |  EUR: {result_eur_ton:.2f}/ton  |  CHF: {result_chf_ton:.2f}/ton", small_style)],
    ]
    ton_table = Table(ton_data, colWidths=[170*mm])
    ton_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#ebf2fa')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#296291')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(ton_table)
    
    # Footer
    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph('Birim Satış Fiyatı Hesaplayıcı - TCMB döviz kurları ile otomatik hesaplama', footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    product = calc.get('product_name', 'hesaplama').replace(' ', '_')
    filename = f"hesaplama_{product}_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """Export all calculations to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    data = request.get_json()
    calculations = data.get('calculations', [])
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hesaplamalar'
    
    # Header styling
    header_fill = PatternFill(start_color='296291', end_color='296291', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=9)
    cell_font = Font(size=9)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    headers = [
        'Tarih', 'Ürün İsmi', 'Bayi', 'Bayi Müşteri', 'Fabrika', 'Nakliye Şehri',
        'Hesaplama Tarihi', 'USD Kur', 'EUR Kur', 'CHF Kur',
        'Nakliye (TL/kg)', 'Nakliye (USD/kg)', 'NTS Maliyeti (TL/kg)', 'Marj (%)',
        'Fiyat (TL/kg)', 'Fiyat (USD/kg)', 'Fiyat (EUR/kg)', 'Fiyat (CHF/kg)',
        'Fiyat (TL/ton)', 'Fiyat (USD/ton)', 'Fiyat (EUR/ton)', 'Fiyat (CHF/ton)'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    factory_labels = {'adana': 'Adana', 'trabzon': 'Trabzon', 'gebze': 'Gebze'}
    alt_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
    
    for idx, c in enumerate(calculations):
        row = idx + 2
        shipping_cost_usd = c.get('shipping_cost_usd', 0)
        if shipping_cost_usd == 0 and c.get('usd_rate', 0) > 0:
            shipping_cost_usd = c.get('shipping_cost', 0) / c.get('usd_rate', 1)
        
        values = [
            c.get('created_at', '')[:16].replace('T', ' '),
            c.get('product_name', ''),
            c.get('dealer', ''),
            c.get('dealer_customer', ''),
            factory_labels.get(c.get('factory', ''), c.get('factory', '')),
            c.get('shipping_city', ''),
            c.get('calculation_date', ''),
            c.get('usd_rate', 0),
            c.get('eur_rate', 0),
            c.get('chf_rate', 0),
            c.get('shipping_cost', 0),
            shipping_cost_usd,
            c.get('nts_cost', 0),
            int(c.get('margin', 0.7) * 100),
            c.get('result_tl', 0),
            c.get('result_usd', 0),
            c.get('result_eur', 0),
            c.get('result_chf', 0),
            c.get('result_tl_ton', c.get('result_tl', 0) * 1000),
            c.get('result_usd_ton', c.get('result_usd', 0) * 1000),
            c.get('result_eur_ton', c.get('result_eur', 0) * 1000),
            c.get('result_chf_ton', c.get('result_chf', 0) * 1000),
        ]
        
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = alt_fill
            # Number formatting
            if col in (8, 9, 10, 12, 16, 17, 18):  # 4 decimal
                cell.number_format = '#,##0.0000'
            elif col in (11, 13, 15, 19, 20, 21, 22):  # 2 decimal
                cell.number_format = '#,##0.00'
    
    # Column widths
    widths = [18, 24, 18, 18, 10, 16, 14, 12, 12, 12, 14, 14, 16, 8, 14, 14, 14, 14, 16, 14, 14, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26)].width = w
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"hesaplamalar_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/api/export/excel/single', methods=['POST'])
def export_single_excel():
    """Export a single calculation to Excel (key-value detail format)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    calc = request.get_json()
    if not calc:
        return jsonify({'error': 'Geçersiz veri'}), 400
    
    factory_labels = {'adana': 'Adana', 'trabzon': 'Trabzon', 'gebze': 'Gebze'}
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hesaplama'
    
    # Styles
    title_font = Font(bold=True, size=14, color='296291')
    section_font = Font(bold=True, size=11, color='296291')
    label_font = Font(bold=True, size=10, color='505050')
    value_font = Font(size=10)
    section_fill = PatternFill(start_color='F0F5FA', end_color='F0F5FA', fill_type='solid')
    
    nts_cost = calc.get('nts_cost', 0)
    margin_val = calc.get('margin', 0.7)
    shipping_cost_usd = calc.get('shipping_cost_usd', 0)
    if shipping_cost_usd == 0 and calc.get('usd_rate', 0) > 0:
        shipping_cost_usd = calc.get('shipping_cost', 0) / calc.get('usd_rate', 1)
    
    result_tl = calc.get('result_tl', 0)
    result_usd = calc.get('result_usd', 0)
    result_eur = calc.get('result_eur', 0)
    result_chf = calc.get('result_chf', 0)
    result_tl_ton = calc.get('result_tl_ton', result_tl * 1000)
    result_usd_ton = calc.get('result_usd_ton', result_usd * 1000)
    result_eur_ton = calc.get('result_eur_ton', result_eur * 1000)
    result_chf_ton = calc.get('result_chf_ton', result_chf * 1000)
    
    data = [
        ('title', 'Birim Satış Fiyatı Hesaplama', ''),
        ('empty', '', ''),
        ('section', 'ÜRÜN VE MÜŞTERİ BİLGİLERİ', ''),
        ('row', 'Ürün İsmi', calc.get('product_name', '')),
        ('row', 'Bayi', calc.get('dealer', '')),
        ('row', 'Bayi Müşteri', calc.get('dealer_customer', '')),
        ('row', 'Sevk Fabrikası', factory_labels.get(calc.get('factory', ''), calc.get('factory', ''))),
        ('row', 'Nakliye Şehri', calc.get('shipping_city', '') or '—'),
        ('empty', '', ''),
        ('section', 'DÖVİZ KURLARI (TCMB SATIŞ)', ''),
        ('row', 'Hesaplama Tarihi', calc.get('calculation_date', '')),
        ('row', 'USD Kur', calc.get('usd_rate', 0)),
        ('row', 'EUR Kur', calc.get('eur_rate', 0)),
        ('row', 'CHF Kur', calc.get('chf_rate', 0)),
        ('empty', '', ''),
        ('section', 'MALİYET VE HESAPLAMA', ''),
        ('row', 'NTS Maliyeti (TL/kg)', nts_cost),
        ('row', 'Nakliye Bedeli (TL/kg)', calc.get('shipping_cost', 0)),
        ('row', 'Nakliye Bedeli (USD/kg)', shipping_cost_usd),
        ('row', 'Hedeflenen Marj', f"%{int(margin_val * 100)}"),
        ('row', 'NTS / Marj (TL/kg)', nts_cost / margin_val if margin_val > 0 else 0),
        ('empty', '', ''),
        ('section', 'SONUÇLAR - KG BAZINDA', ''),
        ('row', 'Birim Fiyat (TL/kg)', result_tl),
        ('row', 'Birim Fiyat (USD/kg)', result_usd),
        ('row', 'Birim Fiyat (EUR/kg)', result_eur),
        ('row', 'Birim Fiyat (CHF/kg)', result_chf),
        ('empty', '', ''),
        ('section', 'SONUÇLAR - TON BAZINDA', ''),
        ('row', 'Birim Fiyat (TL/ton)', result_tl_ton),
        ('row', 'Birim Fiyat (USD/ton)', result_usd_ton),
        ('row', 'Birim Fiyat (EUR/ton)', result_eur_ton),
        ('row', 'Birim Fiyat (CHF/ton)', result_chf_ton),
        ('empty', '', ''),
        ('row', 'Oluşturulma', calc.get('created_at', '')[:16].replace('T', ' ')),
    ]
    
    for idx, (row_type, label, value) in enumerate(data, 1):
        if row_type == 'title':
            cell = ws.cell(row=idx, column=1, value=label)
            cell.font = title_font
        elif row_type == 'section':
            cell = ws.cell(row=idx, column=1, value=label)
            cell.font = section_font
            ws.cell(row=idx, column=1).fill = section_fill
            ws.cell(row=idx, column=2).fill = section_fill
        elif row_type == 'row':
            cell_a = ws.cell(row=idx, column=1, value=label)
            cell_a.font = label_font
            cell_b = ws.cell(row=idx, column=2, value=value)
            cell_b.font = value_font
            if isinstance(value, float):
                if abs(value) >= 100:
                    cell_b.number_format = '#,##0.00'
                else:
                    cell_b.number_format = '#,##0.0000'
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    product = calc.get('product_name', 'hesaplama').replace(' ', '_')
    filename = f"hesaplama_{product}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

# ─── INIT ──────────────────────────────────────────────────────────

init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
